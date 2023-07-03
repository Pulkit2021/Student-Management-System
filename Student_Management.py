from tkinter import *

from datetime import date

from tkinter import filedialog

from tkinter import messagebox

from PIL import Image,ImageTk

import os

from tkinter.ttk import Combobox

import openpyxl, xlrd

from openpyxl import Workbook

import pathlib

background="#06283D"

framebg="#EDEDED"

framefg="#06283D"

 

 

root=Tk()

root.title("Student Registration System")

root.geometry("1250x700+210+100")

root.config(bg=background)

 

 

 

file=pathlib.Path(r'C:\Users\PU20247557\Documents\Student_data.xlsx')

if file.exists():

    pass

else:

    file=Workbook()

    sheet=file.active

    sheet["A1"]="Registartion No."

    sheet["B1"]="Name"

    sheet["C1"]="Class"

    sheet["D1"]="Gender"

    sheet["E1"]="DOB"

    sheet["F1"]="Date of Registration"

    sheet["G1"]="Religion"

    sheet["H1"]="Skill"

    sheet["I1"]="Father's Name"

    sheet["J1"]="Mother's Name"

    sheet["K1"]="Father's Occupation"

    sheet["L1"]="Mother's Occupation"

 

    file.save(r"C:\Users\PU20247557\Documents\Student_data.xlsx")

 

#Exit Window

def Exit():

    root.destroy()

 

def showimage():

    filename=filedialog.askopenfilename(initialdir=os.getcwd(),title="Select Image File",filetypes=(("JPG File","*.jpg"),("PNG File","*.png"),("All Files","*.txt")))

    img=Image.open(filename)

    resized_img=img.resize(150,150)

    photo2=ImageTk.PhotoImage(resized_img)

    lb1.config(image=photo2)

    lb1.image=photo2

def registration_no():

    file=openpyxl.load_workbook(r"C:\Users\PU20247557\Documents\Student_data.xlsx")

    sheet=file.active

    row=sheet.max_row

    max_row_value=sheet.cell(row=row,column=1).value

    try:

        Registration.set(max_row_value+1)

    except:

        Registration.set("1")

 

def Clear():

    Name.set("")

    DOB.set("")

    Religion.set("")

    Skills.set("")

    F_Name.set("")

    M_Name.set("")

    F_Occupation.set("")

    M_Occupation.set("")

    Class.set("Select Class")

    Religion.set("")

 

 

    registration_no()

 

    saveButton.config(state='normal')

 

    img1=PhotoImage(file=r"C:\Users\PU20247557\Downloads\65590.png")

    lb1.config(image=img1)

    lb1.image=img1

def Save():

    R1=Registration.get()

    N1=Name.get()

    C1=Class.get()

    try:

        G1=gender

    except:

        messagebox.showerror("error","Select Gender!")

    D2=DOB.get()

    D1=Date.get()

    Re1=Religion.get()

    S1=Skills.get()

    fathername=F_Name.get()

    mothername=M_Name.get()

    F1=F_Occupation.get()

    M1=M_Occupation.get()

    if N1=="" or R1=="" or C1=="Select Class" or D2=="" or Re1=="" or S1=="" or fathername=="" or mothername=="" or  F1=="" or M1=="":

        messagebox.showerror("error","Some data is missing!")

    else:

        file=openpyxl.load_workbook(r"C:\Users\PU20247557\Documents\Student_data.xlsx")

        sheet=file.active

        sheet.cell(column=1,row=sheet.max_row,value=R1)

        sheet.cell(column=2,row=sheet.max_row,value=N1)

        sheet.cell(column=3,row=sheet.max_row,value=C1)

        sheet.cell(column=4,row=sheet.max_row,value=G1)

        sheet.cell(column=5,row=sheet.max_row,value=D2)

        sheet.cell(column=6,row=sheet.max_row,value=D1)

        sheet.cell(column=7,row=sheet.max_row,value=Re1)

        sheet.cell(column=8,row=sheet.max_row,value=S1)

        sheet.cell(column=9,row=sheet.max_row,value=fathername)

        sheet.cell(column=10,row=sheet.max_row,value=mothername)

        sheet.cell(column=11,row=sheet.max_row,value=F1)

        sheet.cell(column=12,row=sheet.max_row,value=M1)

        file.save(r"C:\Users\PU20247557\Documents\Student_data.xlsx")

 

 

 

# #gender

def selection():

    global gender

    value=radio.get()

    if value==1:

        gender="Male"

    else:

        gender="Female"

        print(gender)

 

#top frames

Label(root,text="Email: registrar.school@gmail.com",width=10,height=3,bg="#f0687c",anchor="e").pack(side=TOP,fill=X)

Label(root,text="Student Registration",width=10,height=2,bg="#c36464",fg="#fff",font="arial 20 bold").pack(side=TOP,fill=X)

 

#search box to update

Search=StringVar()

#Using Entry to display text

Entry(root,textvariable=Search,width=15,bd=2,font="arial 20").place(x=820,y=70)

imageicon3=PhotoImage(file=r"C:\Users\PU20247557\Downloads\9788886.png")

Srch=Button(root,text="Search",compound=LEFT,width=123,bg="#68ddfa",image=imageicon3,font="arial 13 bold")

Srch.place(x=1060,y=50)

 

imageicon4=PhotoImage(file=r"C:\Users\PU20247557\Downloads\65590.png")

Update_button=Button(root,image=imageicon4,bg="#FFF")

Update_button.place(x=110,y=64)

 

#Registration and Date

Label(root,text="Registration No:",font="arial 13",fg=framebg,bg=background).place(x=30,y=150)

Label(root,text="Date:",font="arial 13",fg=framebg,bg=background).place(x=500,y=150)

 

Registration=StringVar()

Date=StringVar()

 

#Registration no

 

reg_entry= Entry(root,textvariable=Registration,width=15,font="arial 10")

reg_entry.place(x=160,y=150)

registration_no()

 

today=date.today()

d1=today.strftime("%d/%m/%y")

date_entry=Entry(root,textvariable=Date,width=15,font="arial 10")

date_entry.place(x=550,y=150)

 

Date.set(d1)

 

#Student Details

obj=LabelFrame(root,text="Student's Details", font=20,bd=2, width=900,bg=framebg,fg="Black",height=250,relief=GROOVE)

obj.place(x=30,y=200)

Label(obj,text="Full Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)

Label(obj,text="Date of Birth:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)

Label(obj,text="Gender:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=150)

 

 

 

Label(obj,text="Class:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)

Label(obj,text="Religion:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)

Label(obj,text="Skills:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=150)

 

Name=StringVar()

name_entry=Entry(obj,textvariable=Name,width=20,font="arial 10")

name_entry.place(x=160,y=50)

 

radio=IntVar()

R1=Radiobutton(obj,text="Male",variable=radio,value=1,bg=framebg,fg=framefg,command=selection)

R1.place(x=150,y=150)

 

R1=Radiobutton(obj,text="Female",variable=radio,value=2,bg=framebg,fg=framefg,command=selection)

R1.place(x=200,y=150)

 

DOB=StringVar()

dob_entry=Entry(obj,textvariable=DOB,width=20,font="arial 10")

dob_entry.place(x=160,y=100)

 

Religion=StringVar()

religion_entry=Entry(obj,textvariable=Religion,width=20,font="arial 10")

religion_entry.place(x=630,y=100)

 

Skills=StringVar()

Skill_entry=Entry(obj,textvariable=Skills,width=20,font="arial 10")

Skill_entry.place(x=630,y=150)

 

Class= Combobox(obj,values=['1','2',"3","4",'5','6',"7","8","9",'10','11','12'],font="Roboto 10",width=17,state="r")

Class.place(x=630,y=50)

Class.set("Select Class")

 

 

 

 

#Parent Details

obj2=LabelFrame(root,text="Parents' Details", font=20,bd=2, width=900,bg=framebg,fg="Black",height=220,relief=GROOVE)

obj2.place(x=30,y=470)

 

Label(obj2,text="Father's Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=50)

Label(obj2,text="Occupation:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=100)

 

Label(obj2,text="Mother's Name:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=50)

Label(obj2,text="Occupation:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=100)

 

F_Name=StringVar()

f_entry=Entry(obj2,textvariable=F_Name,width=20,font="arial 10")

f_entry.place(x=160,y=50)

 

F_Occupation=StringVar()

fo_entry=Entry(obj2,textvariable=F_Occupation,width=20,font="arial 10")

fo_entry.place(x=160,y=100)

 

M_Name=StringVar()

M_entry=Entry(obj2,textvariable=M_Name,width=20,font="arial 10")

M_entry.place(x=630,y=50)

 

M_Occupation=StringVar()

Mo_entry=Entry(obj2,textvariable=M_Occupation,width=20,font="arial 10")

Mo_entry.place(x=630,y=100)

 

#Image

f=Frame(root,bd=3,bg="black",width=200,height=200,relief=GROOVE)

f.place(x=1000,y=150)

 

img=PhotoImage(file=r'C:\Users\PU20247557\Downloads\download.png')

lb1=Label(f,bg="black",image=img)

lb1.place(x=0,y=0)

 

#Button

Button(root,text="Upload",width=19,height=2,font="arial 12 bold",bg="lightblue",command=showimage).place(x=1000,y=370)

saveButton=Button(root,text="Save",width=19,height=2,font="arial 12 bold",bg="lightgreen",command=Save)

saveButton.place(x=1000,y=450)

Button(root,text="Reset",width=19,height=2,font="arial 12 bold",bg="lightpink",command=Clear).place(x=1000,y=530)

Button(root,text="Exit",width=19,height=2,font="arial 12 bold",bg="grey",command=Exit).place(x=1000,y=610)

 

root.mainloop()

 

 